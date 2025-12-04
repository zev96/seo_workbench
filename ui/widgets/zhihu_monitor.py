"""
知乎监测主页面
提供问题监控列表、检测、导出等功能
"""

import json
from datetime import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QHeaderView, QDialog, QFormLayout, QMessageBox,
    QFileDialog, QCheckBox
)
from PyQt6.QtCore import Qt, pyqtSignal, QUrl
from PyQt6.QtGui import QColor, QDesktopServices
from qfluentwidgets import (
    PushButton, LineEdit, SpinBox, MessageBox, ProgressBar,
    FluentIcon, TableWidget, ToolButton, ComboBox
)
from loguru import logger
from sqlalchemy.orm import Session

from ...database.models import ZhihuMonitorTask, ZhihuBrand, ZhihuMonitorHistory
from ...core.zhihu_monitor_worker import ZhihuMonitorWorker
from ...core.zhihu_scheduler import ZhihuScheduler
from ..dialogs.brand_manager_dialog import BrandManagerDialog
from ..dialogs.zhihu_settings_dialog import ZhihuSettingsDialog


class AddTaskDialog(QDialog):
    """添加监控任务对话框"""
    
    def __init__(self, db_session: Session, parent=None):
        super().__init__(parent)
        self.db_session = db_session
        self.setWindowTitle("添加知乎监控")
        self.resize(600, 300)
        
        self._init_ui()
        self._load_brands()
        
    def _init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        
        # 表单区
        form_layout = QFormLayout()
        form_layout.setSpacing(15)
        
        self.url_input = LineEdit()
        self.url_input.setPlaceholderText("https://www.zhihu.com/question/123456789")
        form_layout.addRow("问题链接:", self.url_input)
        
        self.brand_combo = ComboBox()
        form_layout.addRow("监测品牌:", self.brand_combo)
        
        self.range_spin = SpinBox()
        self.range_spin.setRange(5, 50)
        self.range_spin.setValue(20)
        self.range_spin.setSuffix(" 名")
        form_layout.addRow("检测范围:", self.range_spin)
        
        layout.addLayout(form_layout)
        
        # 提示
        tip_label = QLabel(
            "💡 提示：添加后可手动检测或设置定时任务"
        )
        tip_label.setStyleSheet("color: #666; padding: 10px; background: #f5f5f5; border-radius: 5px;")
        layout.addWidget(tip_label)
        
        layout.addStretch()
        
        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.add_btn = PushButton("添加", self, FluentIcon.ADD)
        self.add_btn.clicked.connect(self.accept)
        btn_layout.addWidget(self.add_btn)
        
        self.cancel_btn = PushButton("取消", self)
        self.cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(btn_layout)
        
    def _load_brands(self):
        """加载我方品牌列表"""
        try:
            brands = self.db_session.query(ZhihuBrand).filter(
                ZhihuBrand.brand_type == 'own'
            ).all()
            
            if brands:
                for brand in brands:
                    self.brand_combo.addItem(brand.name)
            else:
                # 如果没有品牌，添加一个提示项
                self.brand_combo.addItem("请先添加品牌")
                self.brand_combo.setEnabled(False)
                
        except Exception as e:
            logger.error(f"加载品牌失败: {e}")
            self.brand_combo.addItem("加载失败")
            self.brand_combo.setEnabled(False)
    
    def get_task_data(self) -> dict:
        """获取任务数据"""
        return {
            'url': self.url_input.text().strip(),
            'brand': self.brand_combo.currentText(),
            'range': self.range_spin.value()
        }


class ZhihuMonitorWidget(QWidget):
    """知乎监测主页面"""
    
    def __init__(self, db_session: Session, parent=None):
        super().__init__(parent)
        self.db_session = db_session
        self.worker = None
        self.scheduler = None
        
        self._init_ui()
        self._load_tasks()
        self._init_scheduler()
        
    def _init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 顶部工具栏
        toolbar_layout = QHBoxLayout()
        
        self.add_btn = PushButton("添加监控", self, FluentIcon.ADD)
        self.add_btn.clicked.connect(self._add_task)
        toolbar_layout.addWidget(self.add_btn)
        
        self.check_btn = PushButton("立即检测", self, FluentIcon.SYNC)
        self.check_btn.clicked.connect(self._start_check)
        toolbar_layout.addWidget(self.check_btn)
        
        self.import_btn = PushButton("导入Excel", self, FluentIcon.FOLDER_ADD)
        self.import_btn.clicked.connect(self._import_excel)
        toolbar_layout.addWidget(self.import_btn)
        
        self.template_btn = PushButton("下载模板", self, FluentIcon.DOWNLOAD)
        self.template_btn.clicked.connect(self._download_template)
        toolbar_layout.addWidget(self.template_btn)
        
        toolbar_layout.addSpacing(10)
        
        # 全选/取消全选按钮
        self.select_all_btn = PushButton("全选", self, FluentIcon.CHECKBOX)
        self.select_all_btn.clicked.connect(lambda: self._on_select_all(Qt.CheckState.Checked.value))
        toolbar_layout.addWidget(self.select_all_btn)
        
        self.deselect_all_btn = PushButton("取消全选", self, FluentIcon.CANCEL)
        self.deselect_all_btn.clicked.connect(lambda: self._on_select_all(Qt.CheckState.Unchecked.value))
        toolbar_layout.addWidget(self.deselect_all_btn)
        
        toolbar_layout.addSpacing(10)
        
        self.batch_delete_btn = PushButton("批量删除", self, FluentIcon.DELETE)
        self.batch_delete_btn.clicked.connect(self._batch_delete)
        toolbar_layout.addWidget(self.batch_delete_btn)
        
        self.export_btn = PushButton("导出报告", self, FluentIcon.DOWNLOAD)
        self.export_btn.clicked.connect(self._export_report)
        toolbar_layout.addWidget(self.export_btn)
        
        toolbar_layout.addSpacing(20)
        
        self.brand_mgr_btn = ToolButton(FluentIcon.TAG, self)
        self.brand_mgr_btn.setToolTip("品牌管理")
        self.brand_mgr_btn.clicked.connect(self._open_brand_manager)
        toolbar_layout.addWidget(self.brand_mgr_btn)
        
        self.settings_btn = ToolButton(FluentIcon.SETTING, self)
        self.settings_btn.setToolTip("设置")
        self.settings_btn.clicked.connect(self._open_settings)
        toolbar_layout.addWidget(self.settings_btn)
        
        toolbar_layout.addStretch()
        
        # 统计信息
        self.stats_label = QLabel("共 0 个监控任务")
        self.stats_label.setStyleSheet("color: #666; font-size: 13px;")
        toolbar_layout.addWidget(self.stats_label)
        
        layout.addLayout(toolbar_layout)
        
        # 进度条（检测时显示）
        self.progress_bar = ProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # 监控列表表格
        self.table = TableWidget()
        self.table.setColumnCount(8)  # 增加一列用于复选框
        self.table.setHorizontalHeaderLabels([
            "☐", "问题链接", "目标品牌", "状态", "排名", "浏览量/关注", "最后更新", "操作"
        ])
        
        # 设置列宽
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)  # 复选框列固定宽度
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # 标题列自适应
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
        
        self.table.setColumnWidth(0, 50)   # 复选框列
        self.table.setColumnWidth(2, 120)  # 目标品牌
        self.table.setColumnWidth(3, 100)  # 状态
        self.table.setColumnWidth(4, 150)  # 排名
        self.table.setColumnWidth(5, 130)  # 浏览量/关注
        self.table.setColumnWidth(6, 150)  # 最后更新
        self.table.setColumnWidth(7, 150)  # 操作
        
        # 双击标题打开链接
        self.table.cellDoubleClicked.connect(self._on_cell_double_clicked)
        
        layout.addWidget(self.table)
        
    def _load_tasks(self):
        """加载监控任务列表"""
        try:
            tasks = self.db_session.query(ZhihuMonitorTask).order_by(
                ZhihuMonitorTask.created_at.asc()
            ).all()
            
            self.table.setRowCount(len(tasks))
            
            for row, task in enumerate(tasks):
                # 第0列：复选框（默认不勾选）
                checkbox = QCheckBox()
                checkbox.setProperty('task_id', task.id)
                checkbox_widget = QWidget()
                checkbox_layout = QHBoxLayout(checkbox_widget)
                checkbox_layout.addWidget(checkbox)
                checkbox_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
                checkbox_layout.setContentsMargins(0, 0, 0, 0)
                self.table.setCellWidget(row, 0, checkbox_widget)
                
                # 第1列：问题链接（始终显示URL，标题作为 tooltip）
                url_item = QTableWidgetItem(task.question_url)
                # 如果有标题，设置为 tooltip；否则显示 URL 作为 tooltip
                tooltip_text = task.question_title if task.question_title else task.question_url
                url_item.setToolTip(tooltip_text)
                url_item.setData(Qt.ItemDataRole.UserRole, task.question_url)
                self.table.setItem(row, 1, url_item)
                
                # 第2列：目标品牌
                brand_item = QTableWidgetItem(task.target_brand)
                brand_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row, 2, brand_item)
                
                # 第3列：状态
                status_text, status_color = self._get_status_display(task)
                status_item = QTableWidgetItem(status_text)
                status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if status_color:
                    status_item.setBackground(QColor(status_color))
                self.table.setItem(row, 3, status_item)
                
                # 第4列：排名
                ranks = task.get_result_list()
                if ranks:
                    rank_text = f"第 {', '.join(map(str, ranks))} 名"
                else:
                    rank_text = "-"
                rank_item = QTableWidgetItem(rank_text)
                rank_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row, 4, rank_item)
                
                # 第5列：浏览量/关注
                views_text = f"{self._format_number(task.total_views)} / {self._format_number(task.total_followers)}"
                views_item = QTableWidgetItem(views_text)
                views_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                views_item.setToolTip(f"浏览量: {task.total_views}\n关注者: {task.total_followers}")
                self.table.setItem(row, 5, views_item)
                
                # 第6列：最后更新
                update_time = task.last_check_at or task.created_at
                time_text = update_time.strftime("%m-%d %H:%M") if update_time else "-"
                time_item = QTableWidgetItem(time_text)
                time_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row, 6, time_item)
                
                # 第7列：操作按钮
                btn_widget = self._create_action_buttons(task.id, task.schedule_enabled == 1)
                self.table.setCellWidget(row, 7, btn_widget)
            
            # 更新统计
            self.stats_label.setText(f"共 {len(tasks)} 个监控任务")
            
            logger.info(f"已加载 {len(tasks)} 个监控任务")
            
        except Exception as e:
            logger.error(f"加载任务列表失败: {e}")
            MessageBox("错误", f"加载失败: {e}", self).exec()
    
    def _get_status_display(self, task: ZhihuMonitorTask) -> tuple:
        """获取状态显示文本和颜色"""
        # 优先检查last_check_at，如果有检测时间说明已检测过
        if task.last_check_at:
            if task.get_result_list():
                return "✅ 在榜", "#d4edda"
            else:
                # 检测过但未找到
                if task.status == 'failed':
                    return "⚠️ 失败", "#fff3cd"
                else:
                    return "❌ 未发现", "#f8d7da"
        else:
            return "⏳ 待检测", "#d1ecf1"
    
    def _format_number(self, num: int) -> str:
        """格式化数字显示"""
        if num >= 10000:
            return f"{num / 10000:.1f}万"
        elif num >= 1000:
            return f"{num / 1000:.1f}k"
        else:
            return str(num)
    
    def _create_action_buttons(self, task_id: int, is_scheduled: bool = False) -> QWidget:
        """创建操作按钮组"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(5, 2, 5, 2)
        layout.setSpacing(5)
        
        # 详情按钮
        detail_btn = QPushButton("📊")
        detail_btn.setToolTip("查看详细分析")
        detail_btn.clicked.connect(lambda: self._show_detail(task_id))
        detail_btn.setFixedWidth(35)
        layout.addWidget(detail_btn)
        
        # 定时任务按钮
        schedule_btn = QPushButton("⏰" if is_scheduled else "⏰")
        schedule_btn.setToolTip("配置定时任务")
        schedule_btn.clicked.connect(lambda: self._config_schedule(task_id))
        schedule_btn.setFixedWidth(35)
        if is_scheduled:
            schedule_btn.setStyleSheet("background-color: #4CAF50; color: white;")
        layout.addWidget(schedule_btn)
        
        # 删除按钮
        delete_btn = QPushButton("🗑️")
        delete_btn.setToolTip("删除监控")
        delete_btn.clicked.connect(lambda: self._delete_task(task_id))
        delete_btn.setFixedWidth(35)
        layout.addWidget(delete_btn)
        
        return widget
    
    def _on_cell_double_clicked(self, row: int, column: int):
        """双击单元格事件"""
        if column == 1:  # 标题列（现在是第1列）
            item = self.table.item(row, 1)
            if item:
                url = item.data(Qt.ItemDataRole.UserRole)
                if url:
                    QDesktopServices.openUrl(QUrl(url))
                    logger.info(f"打开链接: {url}")
    
    def _on_select_all(self, state):
        """全选/取消全选"""
        checked = (state == Qt.CheckState.Checked.value)
        
        for row in range(self.table.rowCount()):
            checkbox_widget = self.table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(checked)
        
        logger.info(f"{'全选' if checked else '取消全选'}所有任务")
    
    def _get_selected_task_ids(self) -> list:
        """获取所有勾选的任务ID"""
        selected_ids = []
        
        for row in range(self.table.rowCount()):
            checkbox_widget = self.table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox and checkbox.isChecked():
                    task_id = checkbox.property('task_id')
                    if task_id:
                        selected_ids.append(task_id)
        
        return selected_ids
    
    def _add_task(self):
        """添加监控任务"""
        try:
            dialog = AddTaskDialog(self.db_session, self)
            
            if dialog.exec():
                task_data = dialog.get_task_data()
                
                if not task_data['url']:
                    MessageBox("提示", "请输入问题链接", self).exec()
                    return
                
                if not task_data['brand'] or task_data['brand'] == "请先添加品牌":
                    MessageBox("提示", "请先在品牌管理中添加我方品牌", self).exec()
                    return
                
                try:
                    # 检查是否已存在
                    exists = self.db_session.query(ZhihuMonitorTask).filter(
                        ZhihuMonitorTask.question_url == task_data['url']
                    ).first()
                    
                    if exists:
                        MessageBox("提示", "该问题已在监控列表中", self).exec()
                        return
                    
                    # 创建新任务
                    new_task = ZhihuMonitorTask(
                        question_url=task_data['url'],
                        target_brand=task_data['brand'],
                        check_range=task_data['range'],
                        status='pending'
                    )
                    
                    self.db_session.add(new_task)
                    self.db_session.commit()
                    
                    logger.success(f"监控任务已添加: {task_data['url']}")
                    MessageBox("成功", "监控任务已添加", self).exec()
                    
                    # 刷新列表
                    self._load_tasks()
                    
                except Exception as e:
                    self.db_session.rollback()
                    logger.error(f"添加任务失败: {e}")
                    MessageBox("错误", f"添加失败: {e}", self).exec()
                    
        except Exception as e:
            logger.error(f"打开添加对话框失败: {e}")
            MessageBox("错误", f"打开对话框失败: {e}", self).exec()
    
    def _start_check(self):
        """开始检测"""
        # 获取勾选的任务ID
        selected_ids = self._get_selected_task_ids()
        
        if not selected_ids:
            MessageBox("提示", "请先勾选要检测的任务", self).exec()
            return
        
        # 获取勾选的任务
        tasks = self.db_session.query(ZhihuMonitorTask).filter(
            ZhihuMonitorTask.id.in_(selected_ids)
        ).all()
        
        if not tasks:
            MessageBox("提示", "暂无可检测的任务", self).exec()
            return
        
        # 获取配置
        settings_dialog = ZhihuSettingsDialog(self.db_session, self)
        config = settings_dialog.get_config_dict()
        
        # ✅ 关键检查：验证 ChromeDriver 路径是否已配置
        chromedriver_path = config.get('chromedriver_path')
        if not chromedriver_path:
            MessageBox(
                "配置错误",
                "未配置 ChromeDriver 路径！\n\n"
                "请先点击右上角【设置】按钮，\n"
                "在 ChromeDriver 配置 区域选择 chromedriver.exe 文件路径。\n\n"
                "下载地址：\n"
                "https://googlechromelabs.github.io/chrome-for-testing/",
                self
            ).exec()
            logger.error("❌ 未配置 ChromeDriver 路径，无法启动检测")
            return
        
        # 验证路径是否存在
        import os
        if not os.path.exists(chromedriver_path):
            MessageBox(
                "路径错误",
                f"ChromeDriver 文件不存在：\n{chromedriver_path}\n\n"
                "请重新配置正确的路径。",
                self
            ).exec()
            logger.error(f"❌ ChromeDriver 文件不存在: {chromedriver_path}")
            return
        
        logger.info(f"✅ ChromeDriver 路径验证通过: {chromedriver_path}")
        
        # 获取品牌关键词
        brands = self.db_session.query(ZhihuBrand).all()
        brand_keywords = [brand.name for brand in brands]
        
        # 准备任务数据
        task_list = []
        for task in tasks:
            task_list.append({
                'id': task.id,
                'url': task.question_url,
                'title': task.question_title,
                'target_brand': task.target_brand,
                'check_range': task.check_range
            })
        
        logger.info("="*60)
        logger.info(f"📋 准备检测 {len(task_list)} 个任务")
        for i, task in enumerate(task_list, 1):
            logger.info(f"  {i}. {task['title'] or task['url'][:50]}")
        logger.info(f"🔧 配置信息:")
        logger.info(f"  - ChromeDriver: {chromedriver_path}")
        logger.info(f"  - 反检测强度: {config.get('anti_detect_level', 'medium')}")
        logger.info(f"  - Cookie: {'已配置' if config.get('cookie') else '未配置'}")
        logger.info("="*60)
        
        # 创建工作线程
        logger.info("🚀 创建工作线程...")
        self.worker = ZhihuMonitorWorker(task_list, config, brand_keywords)
        self.worker.progress_updated.connect(self._on_progress_updated)
        self.worker.task_completed.connect(self._on_task_completed)
        self.worker.task_failed.connect(self._on_task_failed)
        self.worker.all_completed.connect(self._on_all_completed)
        
        # 禁用按钮
        self.check_btn.setEnabled(False)
        self.add_btn.setEnabled(False)
        
        # 显示进度条
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        # 启动线程
        logger.info("🎬 启动检测线程...")
        try:
            self.worker.start()
            logger.success(f"✅ 线程已启动，开始检测 {len(task_list)} 个任务")
        except Exception as e:
            logger.error(f"❌ 启动线程失败: {e}")
            MessageBox("错误", f"启动检测失败:\n{e}", self).exec()
            self.check_btn.setEnabled(True)
            self.add_btn.setEnabled(True)
            self.progress_bar.setVisible(False)
    
    def _on_progress_updated(self, current: int, total: int, message: str):
        """进度更新"""
        progress = int((current / total) * 100)
        self.progress_bar.setValue(progress)
        logger.info(f"进度: {current}/{total} - {message}")
    
    def _on_task_completed(self, task_id: int, result: dict):
        """任务完成"""
        try:
            task = self.db_session.query(ZhihuMonitorTask).filter(
                ZhihuMonitorTask.id == task_id
            ).first()
            
            if task:
                task.question_title = result.get('question_title')
                task.question_detail = result.get('question_detail', '')
                task.total_views = result.get('total_views', 0)
                task.total_followers = result.get('total_followers', 0)
                task.set_result_list(result.get('found_ranks', []))
                task.status = result.get('status', 'success')
                task.last_check_at = datetime.now()
                
                # 保存Top10快照（重要！）
                if 'top10_snapshot' in result:
                    task.set_snapshot(result['top10_snapshot'])
                    logger.info(f"已保存Top10快照，包含 {len(result['top10_snapshot'].get('top10', []))} 条数据")
                
                # 保存历史记录
                history = ZhihuMonitorHistory(
                    task_id=task_id,
                    check_result=json.dumps(result.get('found_ranks', [])),
                    total_views=result.get('total_views', 0),
                    total_followers=result.get('total_followers', 0),
                    snapshot_data=json.dumps(result.get('top10_snapshot', {})) if 'top10_snapshot' in result else None
                )
                self.db_session.add(history)
                
                self.db_session.commit()
                
                logger.success(f"任务 {task_id} 结果已保存")
                
        except Exception as e:
            self.db_session.rollback()
            logger.error(f"保存任务结果失败: {e}")
    
    def _on_task_failed(self, task_id: int, error: str):
        """任务失败"""
        try:
            task = self.db_session.query(ZhihuMonitorTask).filter(
                ZhihuMonitorTask.id == task_id
            ).first()
            
            if task:
                task.status = 'failed'
                task.last_check_at = datetime.now()
                self.db_session.commit()
                
                logger.error(f"任务 {task_id} 失败: {error}")
                
        except Exception as e:
            self.db_session.rollback()
            logger.error(f"更新任务状态失败: {e}")
    
    def _on_all_completed(self):
        """所有任务完成"""
        self.progress_bar.setVisible(False)
        self.check_btn.setEnabled(True)
        self.add_btn.setEnabled(True)
        
        # 刷新列表
        self._load_tasks()
        
        MessageBox("完成", "所有监控任务已检测完成", self).exec()
        logger.success("所有任务检测完成")
    
    def _delete_task(self, task_id: int):
        """删除任务"""
        try:
            task = self.db_session.query(ZhihuMonitorTask).filter(
                ZhihuMonitorTask.id == task_id
            ).first()
            
            if not task:
                return
            
            reply = MessageBox(
                "确认删除",
                f"确定要删除监控任务吗？\n{task.question_title or task.question_url}",
                self
            )
            
            if reply.exec():
                self.db_session.delete(task)
                self.db_session.commit()
                
                logger.success(f"任务已删除: {task_id}")
                
                # 刷新列表
                self._load_tasks()
                
        except Exception as e:
            self.db_session.rollback()
            logger.error(f"删除任务失败: {e}")
            MessageBox("错误", f"删除失败: {e}", self).exec()
    
    def _show_detail(self, task_id: int):
        """显示详情"""
        from ..dialogs.zhihu_detail_dialog import ZhihuDetailDialog
        
        dialog = ZhihuDetailDialog(task_id, self.db_session, self)
        dialog.exec()
    
    def _export_report(self):
        """导出Excel报告"""
        try:
            # 获取所有任务
            tasks = self.db_session.query(ZhihuMonitorTask).order_by(
                ZhihuMonitorTask.created_at.desc()
            ).all()
            
            if not tasks:
                MessageBox("提示", "暂无数据可导出", self).exec()
                return
            
            # 选择保存位置
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "导出知乎监测报告",
                f"知乎监测报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel文件 (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # 导出数据
            self._export_to_excel(tasks, file_path)
            
            MessageBox("成功", f"报告已导出至:\n{file_path}", self).exec()
            logger.success(f"报告已导出: {file_path}")
            
        except Exception as e:
            logger.error(f"导出报告失败: {e}")
            MessageBox("错误", f"导出失败: {e}", self).exec()
    
    def _export_to_excel(self, tasks: list, file_path: str):
        """
        导出数据到Excel
        
        Args:
            tasks: 任务列表
            file_path: 保存路径
        """
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # 准备数据
        data = []
        for task in tasks:
            ranks = task.get_result_list()
            rank_text = f"第 {', '.join(map(str, ranks))} 名" if ranks else "未上榜"
            
            status_map = {
                'success': '✅ 在榜' if ranks else '❌ 未发现',
                'failed': '⚠️ 失败',
                'pending': '⏳ 待检测'
            }
            status_text = status_map.get(task.status, task.status)
            
            update_time = task.last_check_at or task.created_at
            time_text = update_time.strftime("%Y-%m-%d %H:%M") if update_time else "-"
            
            schedule_text = f"✓ {task.schedule_time}" if task.schedule_enabled else "✗"
            
            data.append({
                '问题链接': task.question_url,
                '目标品牌': task.target_brand,
                '检测范围': f"Top {task.check_range}",
                '当前状态': status_text,
                '排名位置': rank_text,
                '浏览量': task.total_views,
                '关注者数': task.total_followers,
                '定时任务': schedule_text,
                '最后更新': time_text
            })
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 导出到Excel
        df.to_excel(file_path, index=False, sheet_name='监控报告')
        
        # 美化Excel
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 设置标题行样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 设置数据行居中
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存
        wb.save(file_path)
        
        logger.info(f"已导出 {len(data)} 条记录到Excel")
    
    def _export_history_report(self, task_id: int):
        """
        导出单个任务的历史趋势报告
        
        Args:
            task_id: 任务ID
        """
        try:
            task = self.db_session.query(ZhihuMonitorTask).filter(
                ZhihuMonitorTask.id == task_id
            ).first()
            
            if not task:
                return
            
            # 获取历史记录
            histories = self.db_session.query(ZhihuMonitorHistory).filter(
                ZhihuMonitorHistory.task_id == task_id
            ).order_by(ZhihuMonitorHistory.check_at.asc()).all()
            
            if not histories:
                MessageBox("提示", "该任务暂无历史记录", self).exec()
                return
            
            # 选择保存位置
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "导出历史趋势报告",
                f"{task.question_title or 'task'}_历史_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel文件 (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # 准备历史数据
            import pandas as pd
            
            data = []
            for history in histories:
                ranks = history.get_result_list()
                rank_text = f"第 {', '.join(map(str, ranks))} 名" if ranks else "未上榜"
                
                data.append({
                    '检测时间': history.check_at.strftime("%Y-%m-%d %H:%M"),
                    '排名位置': rank_text,
                    '浏览量': history.total_views,
                    '关注者数': history.total_followers
                })
            
            df = pd.DataFrame(data)
            
            # 导出
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='历史趋势')
                
                # 美化（简化版）
                workbook = writer.book
                worksheet = writer.sheets['历史趋势']
                
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
            
            MessageBox("成功", f"历史报告已导出至:\n{file_path}", self).exec()
            logger.success(f"历史报告已导出: {file_path}")
            
        except Exception as e:
            logger.error(f"导出历史报告失败: {e}")
            MessageBox("错误", f"导出失败: {e}", self).exec()
    
    def _open_brand_manager(self):
        """打开品牌管理"""
        dialog = BrandManagerDialog(self.db_session, self)
        dialog.exec()
    
    def _open_settings(self):
        """打开设置"""
        dialog = ZhihuSettingsDialog(self.db_session, self)
        dialog.exec()
    
    def _init_scheduler(self):
        """初始化定时调度器"""
        try:
            self.scheduler = ZhihuScheduler(self.db_session)
            self.scheduler.schedule_triggered.connect(self._on_schedule_triggered)
            self.scheduler.task_started.connect(self._on_schedule_started)
            self.scheduler.task_finished.connect(self._on_schedule_finished)
            
            # 启动调度器
            self.scheduler.start()
            
            logger.success("知乎监测调度器已启动")
            
        except Exception as e:
            logger.error(f"初始化调度器失败: {e}")
    
    def _config_schedule(self, task_id: int):
        """配置定时任务"""
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QFormLayout, QTimeEdit, QCheckBox
        from PyQt6.QtCore import QTime
        from qfluentwidgets import PushButton
        
        task = self.db_session.query(ZhihuMonitorTask).filter(
            ZhihuMonitorTask.id == task_id
        ).first()
        
        if not task:
            return
        
        # 创建配置对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("配置定时任务")
        dialog.resize(400, 200)
        
        layout = QVBoxLayout(dialog)
        form_layout = QFormLayout()
        
        # 启用开关
        enable_check = QCheckBox("启用定时任务")
        enable_check.setChecked(task.schedule_enabled == 1)
        form_layout.addRow("状态:", enable_check)
        
        # 时间选择
        time_edit = QTimeEdit()
        if task.schedule_time:
            try:
                hour, minute = map(int, task.schedule_time.split(':'))
                time_edit.setTime(QTime(hour, minute))
            except:
                time_edit.setTime(QTime(10, 0))  # 默认10:00
        else:
            time_edit.setTime(QTime(10, 0))
        
        time_edit.setDisplayFormat("HH:mm")
        form_layout.addRow("执行时间:", time_edit)
        
        layout.addLayout(form_layout)
        
        # 提示
        tip_label = QLabel(
            "💡 提示：每天到达设定时间后自动执行一次检测"
        )
        tip_label.setStyleSheet("color: #666; padding: 10px; background: #f5f5f5; border-radius: 3px;")
        layout.addWidget(tip_label)
        
        layout.addStretch()
        
        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        save_btn = PushButton("保存", dialog)
        cancel_btn = PushButton("取消", dialog)
        
        def save_schedule():
            try:
                task.schedule_enabled = 1 if enable_check.isChecked() else 0
                task.schedule_time = time_edit.time().toString("HH:mm")
                
                self.db_session.commit()
                
                logger.success(f"定时任务已配置: {task.schedule_time}")
                MessageBox("成功", f"定时任务已设置为每天 {task.schedule_time} 执行", self).exec()
                
                dialog.accept()
                
                # 刷新列表
                self._load_tasks()
                
            except Exception as e:
                self.db_session.rollback()
                logger.error(f"保存定时配置失败: {e}")
                MessageBox("错误", f"保存失败: {e}", self).exec()
        
        save_btn.clicked.connect(save_schedule)
        cancel_btn.clicked.connect(dialog.reject)
        
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        
        layout.addLayout(btn_layout)
        
        dialog.exec()
    
    def _on_schedule_triggered(self, time_str: str):
        """定时任务触发"""
        logger.info(f"定时任务触发: {time_str}")
    
    def _on_schedule_started(self, count: int):
        """定时任务开始"""
        logger.info(f"定时任务开始执行: {count} 个任务")
    
    def _on_schedule_finished(self, stats: dict):
        """定时任务完成"""
        success = stats.get('success_count', 0)
        failed = stats.get('failed_count', 0)
        logger.success(f"定时任务完成: 成功 {success}, 失败 {failed}")
        
        # 刷新列表
        self._load_tasks()
    
    def _import_excel(self):
        """导入Excel批量添加任务"""
        try:
            # 选择文件
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "选择Excel文件",
                "",
                "Excel文件 (*.xlsx *.xls)"
            )
            
            if not file_path:
                return
            
            import pandas as pd
            
            # 读取Excel
            df = pd.read_excel(file_path)
            
            # 验证列名
            required_columns = ['问题链接']
            if not all(col in df.columns for col in required_columns):
                MessageBox("错误", "Excel格式错误，必须包含'问题链接'列", self).exec()
                return
            
            # 获取默认品牌
            default_brand = None
            own_brands = self.db_session.query(ZhihuBrand).filter(
                ZhihuBrand.brand_type == 'own'
            ).first()
            if own_brands:
                default_brand = own_brands.name
            else:
                MessageBox("错误", "请先在品牌管理中添加我方品牌", self).exec()
                return
            
            # 批量导入
            added_count = 0
            skipped_count = 0
            
            for index, row in df.iterrows():
                url = str(row['问题链接']).strip()
                
                if not url or not url.startswith('http'):
                    logger.warning(f"跳过无效链接: {url}")
                    skipped_count += 1
                    continue
                
                # 检查是否已存在
                exists = self.db_session.query(ZhihuMonitorTask).filter(
                    ZhihuMonitorTask.question_url == url
                ).first()
                
                if exists:
                    logger.info(f"跳过已存在的链接: {url}")
                    skipped_count += 1
                    continue
                
                # 获取品牌和范围
                brand = row.get('目标品牌', default_brand)
                if pd.isna(brand):
                    brand = default_brand
                
                check_range = row.get('检测范围', 20)
                if pd.isna(check_range):
                    check_range = 20
                else:
                    check_range = int(check_range)
                
                # 创建任务
                new_task = ZhihuMonitorTask(
                    question_url=url,
                    target_brand=str(brand),
                    check_range=check_range,
                    status='pending'
                )
                
                self.db_session.add(new_task)
                added_count += 1
            
            self.db_session.commit()
            
            # 刷新列表
            self._load_tasks()
            
            MessageBox(
                "导入完成",
                f"成功导入 {added_count} 个任务\n跳过 {skipped_count} 个（重复或无效）",
                self
            ).exec()
            
            logger.success(f"Excel导入完成: 新增{added_count}, 跳过{skipped_count}")
            
        except Exception as e:
            self.db_session.rollback()
            logger.error(f"导入Excel失败: {e}")
            MessageBox("错误", f"导入失败: {e}", self).exec()
    
    def _download_template(self):
        """下载Excel导入模板"""
        try:
            import pandas as pd
            
            # 创建模板数据
            template_data = {
                '问题链接': [
                    'https://www.zhihu.com/question/123456789',
                    'https://www.zhihu.com/question/987654321'
                ],
                '目标品牌': ['CEWEY', 'CEWEY'],
                '检测范围': [20, 20]
            }
            
            df = pd.DataFrame(template_data)
            
            # 选择保存位置
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "保存模板",
                f"知乎监测导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel文件 (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # 导出模板
            df.to_excel(file_path, index=False)
            
            MessageBox("成功", f"模板已保存至:\n{file_path}", self).exec()
            logger.success(f"模板已导出: {file_path}")
            
        except Exception as e:
            logger.error(f"导出模板失败: {e}")
            MessageBox("错误", f"导出失败: {e}", self).exec()
    
    def _batch_delete(self):
        """批量删除勾选的任务"""
        try:
            selected_ids = self._get_selected_task_ids()
            
            if not selected_ids:
                MessageBox("提示", "请先勾选要删除的任务", self).exec()
                return
            
            # 确认对话框
            reply = MessageBox(
                "确认删除",
                f"确定要删除 {len(selected_ids)} 个监控任务吗？\n此操作不可恢复！",
                self
            )
            
            if reply.exec():
                # 批量删除
                for task_id in selected_ids:
                    task = self.db_session.query(ZhihuMonitorTask).filter(
                        ZhihuMonitorTask.id == task_id
                    ).first()
                    
                    if task:
                        self.db_session.delete(task)
                
                self.db_session.commit()
                
                logger.success(f"已删除 {len(selected_ids)} 个任务")
                MessageBox("成功", f"已删除 {len(selected_ids)} 个任务", self).exec()
                
                # 刷新列表
                self._load_tasks()
                
        except Exception as e:
            self.db_session.rollback()
            logger.error(f"批量删除失败: {e}")
            MessageBox("错误", f"删除失败: {e}", self).exec()
    
    def closeEvent(self, event):
        """关闭事件"""
        # 停止调度器
        if self.scheduler and self.scheduler.isRunning():
            self.scheduler.stop()
            self.scheduler.wait()
        
        super().closeEvent(event)

